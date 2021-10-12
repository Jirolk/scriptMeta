#! python3     
 
#se importa la libreria openpyxl para leer archivos con extensión xlsx.
from openpyxl import load_workbook
import string
import os

print("\n\n\n\n\n\t\t\t\tProcesando datos...")
#abrimos el documento 
wb = load_workbook("data.xlsx")
#activamos el documento
hoja =wb.active
#asignamos el nombre de la hoja a una varialbe para poder trabajar por ella.
pestana=wb['Sheet1']
#Será necesario tener las columnas maximas y minimas como tambien las filas.
#esto ayudará para poder recorrer la planilla
min_col=hoja.min_column
max_col=hoja.max_column
min_fila=hoja.min_row
max_fila=hoja.max_row
#se crea un abecedario para tomar como referencias a los nombres de las columnas en Excel
cabecera = list(string.ascii_uppercase)
#solo se toma lo necesario.
cabecera_excel=cabecera[0:max_col]
#listas necesarias para el procesamiento
funcionarios=[]
empresa=[]
nombre=[]
dia=[]
cantNombre=[]
contar=0
b=0
os.system("cls")
print("\n\n\n\n\n\t\t\t\tEspere por favor... procesando...")
for i in cabecera_excel:
    for z in range(1,max_fila):
        #recorremos la cabecera para asignar los campos que necesitamos para procesar la informacion
        funcionarios.append(list([pestana[f'B{z}'].value,pestana[f'F{z}'].value,pestana[f'P{z}'].value]))
        #asignamos en variables específicas para realizar una mejor busqueda.
        empresa.append(pestana[f'F{z}'].value)
        nombre.append(pestana[f'B{z}'].value)
        dia.append(pestana[f'P{z}'].value)
#Se eliminan las filas sobrantes de la lista funcionarios
funcionarios.remove(funcionarios[0])
funcionarios.remove(funcionarios[1])
funcionarios.remove(funcionarios[2])
 
#set ayuda a obtener los valores unicos en una lista
empresa=list(set(empresa))
empresa.remove('EMPRESA')
empresa.remove(None)
# ordena de forma alfabetica
empresa.sort()
nombre=list(set(nombre))
nombre.remove('NOME')
nombre.remove(None)

dia=list(set(dia))
dia.remove(None)
dia.remove('DT_LEITORA - Dia')
#variable para numerar empresas
z=0

totalDia=0
#limpia la pantalla
os.system("cls")
for d in dia:
    print("\n  :::::::::::::::::::::::::::::Dia:::::::::::::::::::::::::::::")
    print("  ____________________________",d,"_____________________________")
    for e in empresa:
        for x in range(0,len(funcionarios)):
            if d==funcionarios[x][2]:
                if e ==funcionarios[x][1]:
                    cantNombre.append(funcionarios[x][0])
        cantNombre=(list(set(cantNombre)))    
        contar=len(cantNombre)
        # print(cantNombre)
        cantNombre=[]
        if contar !=0:
            z+=1
            print("\n\t",z,") ",e,"accesos: ",contar)
            totalDia+=contar
           
        elif contar==0:
          b+=1  
        contar=0
    nrEmpre=len(empresa)-b
    print("\n\n  ----------------------------------------------------")
    print("\n    Empresas Presentes en el dia: ", d, " => ",nrEmpre,"\n")
    print("         Total de Accesos: ",totalDia)
    print("  ----------------------------------------------------")
    totalDia=0
    b=0
    z=0
input("Oprima la tecla enter para salir")
